from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
import smtplib
from email.mime.image import MIMEImage
import imghdr
import os
from dotenv import load_dotenv


# send the Email
def sendMessage(subject, body, receipient_email, receipient_name, image):
    load_dotenv()
    sendcount = 0
    sender = os.getenv('EMAIL')
    password = os.getenv('PASSWORD')
    sent = 0
    failedTo = []

    print(
        f'Receiver name is: {receipient_name}, Receiver Email is: {receipient_email}')

    sendTo = receipient_email
    message =  MIMEMultipart("related")
    message['Subject'] = subject
    message['From'] = sender
    message['To'] = sendTo
    message.attach(MIMEText(body))
    sendcount += 1

    with open(f'Certificates/{receipient_name}.png', 'rb') as f:
        img_attachment = f.read()
        file_type = imghdr.what(f.name)
        file_name = f.name
        img = MIMEImage(img_attachment, _subtype=file_type)

    message.attach(img)
    img.add_header("Content-ID", "<{}>".format(file_name))

    # adding image attachment to the email_message object
    # message.add_attachment(img_attachment, maintype='image',
    #                        subtype=file_type, filename=file_name)
    # msg.add_header('Content-Disposition', 'attachment', filename=file_name)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            print('Logging in...')
            smtp.login(sender, password)
            print('Sending...')
            smtp.send_message(message)
            print(f'sent to {receipient_name}')

            sent = sent+1
    except Exception as e:
        print(f'failed with error {e}')
        failedTo.append(sendTo)


def advisory(receipient_email, receipient_name, image):
    body = open('body.txt')
    bodyString = body.read()
    subject = open('subject.txt')
    subjectString = subject.read()
    
    subject = subjectString
    print(body.read())
    body = bodyString 

    sendMessage(subject=subject, body=body, receipient_email=receipient_email,
                receipient_name=receipient_name, image=image)


wb = load_workbook('names.xlsx')
ws = wb.active
list_of_names = []
list_of_emails = []
directory = {}
# gets the length of the rows of the file to limit the iteration below
row_count = ws.max_row

# Check if Certificate directory exists and create it if it does not
hasCertificateDir = False
currentpath = os.getcwd()
temp = os.scandir(currentpath)
for entry in temp:
    if entry.is_dir():
        if entry.name == "Certificates":
            hasCertificateDir = True

if (hasCertificateDir == False):
    print('Cannot find Certificate directory. Creating it now.')
    os.mkdir(currentpath + "/Certificates")

# iterates over all the rows and returns a tuple
for row in ws.iter_rows(min_row=2, max_row=row_count):
    # open the certificate template
    image = Image.open('participants.png')
    draw = ImageDraw.Draw(image)

    # choose font and input the Path as well as size of it
    font = ImageFont.FreeTypeFont(
       'opensans.ttf', size=60)

    # draw the name of the participant to the template based on the x and y axis of image
    draw.text(xy=(725, 680), text='{}'.format(
        row[1].value), fill=(255, 255, 255), font=font)
    image.save('Certificates/{}.png'.format(row[1].value))

    # picks a new image file that matches the name of the
    imageSaved = Image.open(f'Certificates/{row[1].value}.png')

    print(
        f"Email is: {row[0].value}, Name is: {row[1].value}\n")

    advisory(receipient_email=row[0].value,
             receipient_name=row[1].value, image=imageSaved)
